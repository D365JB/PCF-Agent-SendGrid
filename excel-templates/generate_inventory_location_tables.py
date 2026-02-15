from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


def add_table(ws, table_name: str, headers: list[str], rows: list[list[object]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append(row)

    last_row = 1 + len(rows)
    last_col = len(headers)
    ref = f"A1:{get_column_letter(last_col)}{last_row}"

    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

    for idx, header in enumerate(headers, start=1):
        width = min(max(len(str(header)) + 2, 12), 36)
        ws.column_dimensions[get_column_letter(idx)].width = width


def main() -> None:
    out_path = Path(__file__).resolve().parent / "Inventory_Location_Tables_TEMPLATE.xlsx"

    wb = Workbook()
    wb.remove(wb.active)

    # Locations
    ws = wb.create_sheet("Locations")
    locations_headers = [
        "LocationId",
        "Name",
        "Type",
        "Address1",
        "City",
        "State",
        "PostalCode",
        "Country",
        "TimeZone",
        "Latitude",
        "Longitude",
        "Email",
        "Phone",
        "IsActive",
        "Notes",
    ]
    locations_rows = [
        [
            "DC-DAL-01",
            "Dallas Distribution Center",
            "DC",
            "123 Example Rd",
            "Dallas",
            "TX",
            "75201",
            "US",
            "America/Chicago",
            32.7767,
            -96.7970,
            "dal-dc@example.com",
            "+1-555-0101",
            True,
            "",
        ],
        [
            "3PL-SEA-01",
            "Seattle 3PL",
            "3PL",
            "500 Demo Ave",
            "Seattle",
            "WA",
            "98101",
            "US",
            "America/Los_Angeles",
            47.6062,
            -122.3321,
            "sea-3pl@example.com",
            "+1-555-0102",
            True,
            "",
        ],
    ]
    add_table(ws, "tblLocations", locations_headers, locations_rows)

    # Inventory
    ws = wb.create_sheet("Inventory")
    inventory_headers = [
        "InventoryId",
        "SKU",
        "LocationId",
        "LotNumber",
        "SerialNumber",
        "Condition",
        "UOM",
        "OnHandQty",
        "ReservedQty",
        "AvailableQty",
        "InboundQty",
        "ATPDate",
        "ExpirationDate",
        "LastUpdatedAt",
        "Notes",
    ]
    inventory_rows = [
        [
            "INV-0001",
            "SKU-00029",
            "DC-DAL-01",
            "LOT-2026-02-A",
            "",
            "NEW",
            "EA",
            120,
            20,
            100,
            50,
            "2026-02-18",
            "",
            "2026-02-15T00:00:00Z",
            "",
        ],
        [
            "INV-0002",
            "SKU-00097",
            "3PL-SEA-01",
            "",
            "SER-XYZ-000097-01",
            "NEW",
            "EA",
            1,
            0,
            1,
            0,
            "",
            "",
            "2026-02-15T00:00:00Z",
            "Serialized example",
        ],
    ]
    add_table(ws, "tblInventory", inventory_headers, inventory_rows)

    # OrderAllocations (optional but useful for 'where is my order allocated')
    ws = wb.create_sheet("OrderAllocations")
    allocations_headers = [
        "AllocationId",
        "OrderNumber",
        "LineNumber",
        "SKU",
        "LocationId",
        "AllocatedQty",
        "AllocationStatus",
        "PromiseDate",
        "LotNumber",
        "SerialNumber",
        "LastUpdatedAt",
        "Notes",
    ]
    allocations_rows = [
        [
            "ALLOC-6600000680-1",
            "6600000680",
            1,
            "SKU-00029",
            "DC-DAL-01",
            5,
            "Allocated",
            "2026-02-18",
            "LOT-2026-02-A",
            "",
            "2026-02-15T00:00:00Z",
            "",
        ]
    ]
    add_table(ws, "tblOrderAllocations", allocations_headers, allocations_rows)

    meta = wb.create_sheet("README")
    meta["A1"].value = "Inventory + Location tables (copy into master workbook)"
    meta["A2"].value = "This workbook is intended to be copied into SAP_Order_Simulator_TEMPLATE.xlsx (or your own master workbook)."
    meta["A3"].value = "After copying sheets, ensure the Excel Tables remain intact (Insert > Table) and names are preserved."
    meta["A4"].value = "Tables included: Locations, Inventory, OrderAllocations (optional)."
    meta["A5"].value = f"Generated: {datetime.now(timezone.utc).isoformat()}"

    wb.save(out_path)
    print(f"Wrote {out_path}")


if __name__ == "__main__":
    main()
