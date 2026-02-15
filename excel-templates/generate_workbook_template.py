from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


def add_table(ws, table_name: str, headers: list[str], rows: list[list[object]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append(row)

    # openpyxl tables need a ref range that includes header row.
    last_row = 1 + len(rows)
    last_col = len(headers)
    ref = f"A1:{get_column_letter(last_col)}{last_row}"

    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

    # Basic sizing
    for idx, header in enumerate(headers, start=1):
        width = min(max(len(str(header)) + 2, 12), 36)
        ws.column_dimensions[get_column_letter(idx)].width = width


def main() -> None:
    out_path = Path(__file__).resolve().parent / "SAP_Order_Simulator_TEMPLATE.xlsx"

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Orders
    ws = wb.create_sheet("Orders")
    orders_headers = [
        "OrderNumber",
        "OrderType",
        "SalesOrg",
        "SoldTo",
        "SoldToName",
        "ShipTo",
        "ShipToCountry",
        "PONumber",
        "CreatedDate",
        "CustomerReqDelDate",
        "EstShipDate",
        "ShipVia",
        "ShipComplete",
        "DeliveryBlock",
        "BillingBlock",
        "CreditHold",
        "Status",
        "TotalOrderQty",
        "TotalDeliveredQty",
        "TotalOpenQty",
        "Currency",
        "NetValue",
        "LastUpdated",
        "UpdatedBy",
    ]
    orders_rows = [
        [
            "6600000680",
            "OR",
            "3000",
            "C000001",
            "Marriott",
            "S000001",
            "SG",
            "PO-71796144",
            "2026-02-12",
            "2026-02-14",
            "2026-02-26",
            "LTL",
            "N",
            "CREDIT",
            "",
            "",
            "Shipped",
            170,
            70,
            100,
            "USD",
            3678.87,
            "2026-02-11 00:00",
            "<enter name>",
        ]
    ]
    add_table(ws, "tblOrders", orders_headers, orders_rows)

    # OrderLines
    ws = wb.create_sheet("OrderLines")
    orderlines_headers = [
        "LineKey",
        "OrderNumber",
        "LineNumber",
        "SKU",
        "SKUDescription",
        "OrderQty",
        "DeliveredQty",
        "OpenQty",
        "LineStatus",
        "Plant",
        "ShipFromCountry",
        "ShipToCountry",
        "UnitPrice",
        "LineNet",
        "LastUpdated",
    ]
    orderlines_rows = [
        [
            "6600000680-1",
            "6600000680",
            1,
            "SKU-00029",
            "MRO item 029",
            16,
            11,
            5,
            "Open",
            "1000",
            "US",
            "SG",
            18.50,
            296.00,
            "2026-02-11 00:00",
        ],
        [
            "6600000680-2",
            "6600000680",
            2,
            "SKU-00097",
            "Industrial item 097",
            38,
            33,
            5,
            "Open",
            "1000",
            "US",
            "SG",
            34.15,
            1297.70,
            "2026-02-11 00:00",
        ],
    ]
    add_table(ws, "tblOrderLines", orderlines_headers, orderlines_rows)

    # Customers
    ws = wb.create_sheet("Customers")
    customers_headers = [
        "CustomerID",
        "CustomerName",
        "Region",
        "Currency",
        "ShipToID",
        "ShipToCountry",
        "Email",
        "Phone",
    ]
    customers_rows = [["C000001", "Marriott", "APAC", "USD", "S000001", "SG", "demo@example.com", "+1-555-0100"]]
    add_table(ws, "tblCustomers", customers_headers, customers_rows)

    # Products
    ws = wb.create_sheet("Products")
    products_headers = ["SKU", "Description", "ProductFamily", "BaseUnitPrice", "Currency"]
    products_rows = [
        ["SKU-00029", "MRO item 029", "MRO", 18.50, "USD"],
        ["SKU-00097", "Industrial item 097", "Industrial", 34.15, "USD"],
    ]
    add_table(ws, "tblProducts", products_headers, products_rows)

    # Shipments
    ws = wb.create_sheet("Shipments")
    shipments_headers = [
        "ShipmentId",
        "OrderNumber",
        "Mode",
        "Carrier",
        "TrackingNumber",
        "PlannedShipDate",
        "PlannedDeliveryDate",
        "SAPShipmentId",
        "Origin",
        "Destination",
        "CustomerEmailOverride",
        "LastEvaluatedAt",
        "PredictedDeliveryDate",
        "PredictedDelayHours",
        "IsRunningLate",
        "LastMilestoneCode",
        "LastMilestoneAt",
    ]
    shipments_rows = [
        [
            "SHP-6600000680-01",
            "6600000680",
            "parcel",
            "UPS",
            "1Z999AA10123456784",
            "2026-02-15T08:00:00Z",
            "2026-02-22T17:00:00Z",
            80001234,
            "Dallas",
            "Redmond",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ]
    ]
    add_table(ws, "tblShipments", shipments_headers, shipments_rows)

    # ShipmentEvents
    ws = wb.create_sheet("ShipmentEvents")
    shipmentevents_headers = [
        "ShipmentId",
        "TrackingNumber",
        "Source",
        "EventTime",
        "EventCode",
        "Status",
        "EventDescription",
        "Location",
        "EstimatedDeliveryDate",
        "Mode",
    ]
    shipmentevents_rows = [
        [
            "SHP-6600000680-01",
            "1Z999AA10123456784",
            "Carrier",
            "2026-02-15T10:15:00Z",
            "PU",
            "Picked up",
            "Picked up by carrier",
            "Dallas",
            "2026-02-22T17:00:00Z",
            "parcel",
        ],
        [
            "SHP-6600000680-01",
            "1Z999AA10123456784",
            "Carrier",
            "2026-02-16T03:40:00Z",
            "IT",
            "In transit",
            "Departed facility",
            "Dallas",
            "2026-02-22T17:00:00Z",
            "parcel",
        ],
        [
            "SHP-6600000680-01",
            "1Z999AA10123456784",
            "Carrier",
            "2026-02-18T18:05:00Z",
            "AR",
            "Arrived",
            "Arrived at hub",
            "Seattle",
            "2026-02-22T17:00:00Z",
            "parcel",
        ],
    ]
    add_table(ws, "tblShipmentEvents", shipmentevents_headers, shipmentevents_rows)

    # Locations (inventory lookup support)
    ws = wb.create_sheet("Locations")
    locations_headers = [
        "LocationId",
        "Name",
        "Type",
        "Address1",
        "City",
        "State",
        "PostalCode",
        "Country",
        "TimeZone",
        "Latitude",
        "Longitude",
        "Email",
        "Phone",
        "IsActive",
        "Notes",
    ]
    locations_rows = [
        [
            "DC-DAL-01",
            "Dallas Distribution Center",
            "DC",
            "123 Example Rd",
            "Dallas",
            "TX",
            "75201",
            "US",
            "America/Chicago",
            32.7767,
            -96.7970,
            "dal-dc@example.com",
            "+1-555-0101",
            True,
            "",
        ],
        [
            "3PL-SEA-01",
            "Seattle 3PL",
            "3PL",
            "500 Demo Ave",
            "Seattle",
            "WA",
            "98101",
            "US",
            "America/Los_Angeles",
            47.6062,
            -122.3321,
            "sea-3pl@example.com",
            "+1-555-0102",
            True,
            "",
        ],
    ]
    add_table(ws, "tblLocations", locations_headers, locations_rows)

    # Inventory (availability + where-to-get prompts)
    ws = wb.create_sheet("Inventory")
    inventory_headers = [
        "InventoryId",
        "SKU",
        "LocationId",
        "LotNumber",
        "SerialNumber",
        "Condition",
        "UOM",
        "OnHandQty",
        "ReservedQty",
        "AvailableQty",
        "InboundQty",
        "ATPDate",
        "ExpirationDate",
        "LastUpdatedAt",
        "Notes",
    ]
    inventory_rows = [
        [
            "MIL-INV-1002",
            "SKU-00029",
            "DC-DAL-01",
            "LOT-2026-02-A",
            "",
            "GOOD",
            "EA",
            120,
            20,
            100,
            50,
            "2026-02-18",
            "",
            "2026-02-15T00:00:00Z",
            "Example inventory row for availability prompts",
        ],
        [
            "MIL-INV-1003",
            "SKU-00097",
            "3PL-SEA-01",
            "",
            "SER-XYZ-000097-01",
            "GOOD",
            "EA",
            1,
            0,
            1,
            0,
            "",
            "",
            "2026-02-15T00:00:00Z",
            "Serialized example",
        ],
    ]
    add_table(ws, "tblInventory", inventory_headers, inventory_rows)

    # OrderAllocations (links orders/lines -> inventory locations)
    ws = wb.create_sheet("OrderAllocations")
    allocations_headers = [
        "AllocationId",
        "OrderNumber",
        "LineNumber",
        "SKU",
        "LocationId",
        "AllocatedQty",
        "AllocationStatus",
        "PromiseDate",
        "LotNumber",
        "SerialNumber",
        "LastUpdatedAt",
        "Notes",
    ]
    allocations_rows = [
        [
            "ALLOC-6600000680-1",
            "6600000680",
            1,
            "SKU-00029",
            "DC-DAL-01",
            5,
            "Allocated",
            "2026-02-18",
            "LOT-2026-02-A",
            "",
            "2026-02-15T00:00:00Z",
            "",
        ]
    ]
    add_table(ws, "tblOrderAllocations", allocations_headers, allocations_rows)

    # Make it obvious this is a template
    meta = wb.create_sheet("README")
    meta["A1"].value = "Template workbook for PCFCopilot demo"
    meta["A2"].value = "This file contains sanitized sample data and required Excel Tables. Upload to OneDrive/SharePoint and configure GRAPH_DRIVE_ID/GRAPH_ITEM_ID + table IDs/names."
    meta["A3"].value = "Tables included: Orders, OrderLines, Customers, Products, Shipments, ShipmentEvents, Locations, Inventory, OrderAllocations."
    meta["A4"].value = f"Generated: {datetime.now(timezone.utc).isoformat()}"

    wb.save(out_path)
    print(f"Wrote {out_path}")


if __name__ == "__main__":
    main()
